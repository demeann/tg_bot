# bot.py
from aiogram.client.default import DefaultBotProperties
from aiogram.types import User, FSInputFile, Message, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
import asyncio
import os
import re
import time
from dataclasses import dataclass
from typing import List, Optional, Tuple
from pathlib import Path

import aiosqlite
from aiogram import Bot, Dispatcher, Router
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from config import settings
import dns.resolver
import idna
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd

# ---------- Пути/файлы и доступ ----------
BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "data" / "emails.db"
EXCEL_PATH = BASE_DIR / "emails.xlsx"
ASSETS_DIR = BASE_DIR / "assets"
INVITE_SECRET = os.getenv("INVITE_SECRET", "")

INTRO_PHOTO = ASSETS_DIR / "intro.png"
FINAL_PHOTO = ASSETS_DIR / "final.png"

# ---------- Контент квиза ----------
@dataclass
class QuizStep:
    photo_url: Optional[str]
    photo_path: Optional[str]
    photo_file_id: Optional[str]
    caption: str
    options: List[str]
    correct: str
    feedback_ok: str
    feedback_bad: str

def kb(options: List[str]) -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=o)] for o in options],
        resize_keyboard=True
    )

INTRO_TEXT = (
    "<b>Дорогой коллега!👋</b>\n\n"
    "Мы запускаем лимитированное предложение: <b>PHILIP MORRIS AROMA TROPIC MIX</b> "
    "и приготовили для тебя увлекательный квиз!\n\n"
    "Ты можешь ответить на несколько вопросов о продукте, проверить свои знания "
    "и получить возможность выиграть призы.\n\n"
    "<b>Чтобы начать, введи свою почту:</b>"
)

STEPS: List[QuizStep] = [
    QuizStep(
        photo_url=None, photo_path=str(ASSETS_DIR / "q1.png"), photo_file_id=None,
        caption="<b>Привет!👋</b>\n\n<b>Итак, начнем!</b> Что послужило направлением при создании дизайна пачки "
                "для PHILIP MORRIS AROMA TROPIC MIX?",
        options=["Ночной город", "Зимний лес", "Тропический пляж"],
        correct="Тропический пляж",
        feedback_ok="Ты отлично уловил суть! Именно тропический пляж на закате изображён на упаковке.",
        feedback_bad="Не совсем то. Тропический пляж на закате изображён на упаковке."
    ),
    QuizStep(
        photo_url=None, photo_path=str(ASSETS_DIR / "q2.png"), photo_file_id=None,
        caption="<b>Привет!👋</b>\n\nНа главном изображении продукта есть коммуникация, которая подчеркивает "
                "сочетание двух элементов: аромата тропических фруктов и вкуса ягодной капсулы.\n\n"
                "<b>Какое слово пропущено?</b>",
        options=["Микс", "Ансамбль", "Союз"],
        correct="Микс",
        feedback_ok=('Правильно!\n"Яркий микс вкуса и аромата" — именно так звучит слоган. '
                     'Слово "микс" показывает сочетание аромата тропических фруктов и капсулы со вкусом ягод.'),
        feedback_bad=('Неправильно.\nСлоган звучит как "Яркий микс вкуса и аромата" и подчёркивает сочетание двух '
                      'компонентов: аромата тропических фруктов и капсулы со вкусом ягод.')
    ),
    QuizStep(
        photo_url=None, photo_path=str(ASSETS_DIR / "q3.png"), photo_file_id=None,
        caption="<b>Привет!👋</b>\n\nПредлагаем внимательно изучить пачку PHILIP MORRIS AROMA TROPIC MIX. "
                "Какой важный элемент дизайна здесь отсутствует?",
        options=["Пальмовая ветвь", "Морская волна", "Тропический цветок"],
        correct="Пальмовая ветвь",
        feedback_ok="Точно подмечено! Пальмовая ветвь — ключевой элемент дизайна пачки PHILIP MORRIS AROMA TROPIC MIX.",
        feedback_bad=("Почти угадал! Верный ответ — пальмовая ветвь. Этот элемент является частью общего дизайна "
                      "пачки PHILIP MORRIS AROMA TROPIC MIX.")
    ),
    QuizStep(
        photo_url=None, photo_path=str(ASSETS_DIR / "q4.png"), photo_file_id=None,
        caption="<b>Привет!👋</b>\n\nТеперь о формате. На картинке ты можешь увидеть три разных формата. "
                "Как ты считаешь, в каком из них выполнен PHILIP MORRIS AROMA TROPIC MIX?",
        options=["Стандартный", "Супертонкий", "Компактный"],
        correct="Компактный",
        feedback_ok="Ты прав! PHILIP MORRIS AROMA TROPIC MIX выпускается в компактном формате!",
        feedback_bad="Не совсем. PHILIP MORRIS AROMA TROPIC MIX выпускается в компактном формате."
    ),
]

FINAL_TEXT = (
    "<b>Это был последний вопрос!</b>\n\n"
    "<b>Спасибо за участие в нашем тропическом квизе!</b>\n\n"
    "Ты отлично справился и теперь знаешь все преимущества нашего предложения:\n"
    "• Компактный формат\n"
    "• Аромат кокоса и ананаса\n"
    "• Капсула со вкусом сочных ягод\n"
    "• Приятный вкус на фильтре\n\n"
    "<b>Спасибо за участие!</b>\n\n"
    "Результаты будут объявлены до xx.xx.xxxx"
)

# ---------- Состояния ----------
class Quiz(StatesGroup):
    q = State()
    score = State()

class CollectEmail(StatesGroup):
    wait_email = State()

router = Router(name="quiz")

# ---------- Вспомогательные функции ----------
EMAIL_RE = re.compile(r"^(?P<local>[^@\s]{1,64})@(?P<domain>[A-Za-z0-9.\-\u0080-\uFFFF]{1,255})$")

async def dedupe_user_ids():
    """Удаляет дубликаты по user_id, оставляя самую раннюю запись (MIN(id))."""
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("""
            SELECT user_id, MIN(id) as keep_id
            FROM emails
            GROUP BY user_id
            HAVING COUNT(*) > 1
        """)
        rows = await cur.fetchall()
        for user_id, keep_id in rows:
            await db.execute("DELETE FROM emails WHERE user_id = ? AND id <> ?", (user_id, keep_id))
        if rows:
            await db.commit()

async def ensure_db():
    (BASE_DIR / "data").mkdir(parents=True, exist_ok=True)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS emails(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                username TEXT,
                first_name TEXT,
                last_name TEXT,
                email TEXT NOT NULL UNIQUE,
                domain TEXT NOT NULL,
                mx_ok INTEGER NOT NULL,
                created_at INTEGER NOT NULL
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS access_users(
                user_id INTEGER PRIMARY KEY,
                granted_at INTEGER NOT NULL
            )
        """)
        await db.commit()

    await dedupe_user_ids()
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("CREATE UNIQUE INDEX IF NOT EXISTS uniq_emails_user ON emails(user_id)")
            await db.commit()
    except aiosqlite.IntegrityError:
        await dedupe_user_ids()
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("CREATE UNIQUE INDEX IF NOT EXISTS uniq_emails_user ON emails(user_id)")
            await db.commit()

async def has_participated(user_id: int) -> bool:
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT 1 FROM emails WHERE user_id = ? LIMIT 1", (user_id,))
        return (await cur.fetchone()) is not None

async def get_user_email(user_id: int) -> Optional[str]:
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT email FROM emails WHERE user_id = ? LIMIT 1", (user_id,))
        row = await cur.fetchone()
        return row[0] if row else None

def domain_to_ascii(domain: str) -> str:
    labels = domain.strip().strip(".").split(".")
    ascii_labels = [idna.encode(lbl).decode("ascii") for lbl in labels if lbl]
    return ".".join(ascii_labels)

def check_mx(domain_ascii: str) -> bool:
    try:
        resolver = dns.resolver.Resolver()
        resolver.timeout = 3.0
        resolver.lifetime = 5.0
        answers = resolver.resolve(domain_ascii, "MX")
        return len(answers) > 0
    except Exception:
        return False

def parse_email(email: str) -> Optional[Tuple[str, str]]:
    m = EMAIL_RE.match(email.strip())
    if not m:
        return None
    local = m.group("local")
    domain = m.group("domain")
    if len(local) > 64 or len(domain) > 255:
        return None
    return local, domain

# Строгая проверка формата e-mail (до MX)
EMAIL_FORMAT_RE = re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')
def is_valid_email(email: str) -> bool:
    return bool(email and isinstance(email, str) and EMAIL_FORMAT_RE.match(email.strip()))

async def save_email_row(user: User, email: str, domain: str, mx_ok: bool):
    """Сохраняет email в БД с проверкой уникальности"""
    ts = int(time.time())
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("""
                INSERT INTO emails(user_id, username, first_name, last_name, email, domain, mx_ok, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (user.id, user.username, user.first_name, user.last_name, email, domain, int(mx_ok), ts))
            await db.commit()
        return True, "Успешно сохранено"
    except aiosqlite.IntegrityError as e:
        msg = str(e).lower()
        if "emails.email" in msg or "unique constraint failed: emails.email" in msg:
            return False, "Email уже существует в базе данных"
        if "uniq_emails_user" in msg or "emails.user_id" in msg:
            return False, "Вы уже проходили квиз (запись по вашему аккаунту существует)"
        return False, "Нарушение уникальности (повторная запись)"
    except Exception as e:
        return False, f"Ошибка сохранения: {str(e)}"

def append_excel_row(email: str, domain: str, mx_ok: bool, user: User) -> None:
    """Добавляет строку в Excel файл с проверкой дубликатов"""
    if not EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Emails"
        ws.append(["user_id", "username", "first_name", "last_name", "email", "domain", "mx_ok", "created_at"])
        wb.save(str(EXCEL_PATH))

    try:
        df = pd.read_excel(str(EXCEL_PATH))
        if 'email' in df.columns:
            existing_emails = df['email'].astype(str).str.lower().values
            if email.lower() in existing_emails:
                print(f"Предупреждение: Email {email} уже существует в Excel файле")
                return
    except Exception as e:
        print(f"Ошибка при проверке Excel на дубликаты: {e}")

    wb = load_workbook(str(EXCEL_PATH))
    ws = wb.active
    ws.append([user.id, user.username, user.first_name, user.last_name,
               email, domain, int(mx_ok), int(time.time())])
    for idx, _ in enumerate(ws[1], start=1):
        ws.column_dimensions[get_column_letter(idx)].auto_size = True
    try:
        wb.save(str(EXCEL_PATH))
    except PermissionError:
        print("Excel-файл открыт. Закройте его и выполните /export для синхронизации.")

async def export_excel_from_db() -> int:
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("""
            SELECT user_id, username, first_name, last_name, email, domain, mx_ok, created_at
            FROM emails ORDER BY id
        """)
        rows = await cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Emails"
    ws.append(["user_id", "username", "first_name", "last_name", "email", "domain", "mx_ok", "created_at"])
    for r in rows:
        ws.append(list(r))
    wb.save(str(EXCEL_PATH))
    return len(rows)

# ---------- Отправка шагов ----------
async def send_step(message: Message, step: QuizStep):
    cap = step.caption
    markup = kb(step.options)
    if step.photo_path and Path(step.photo_path).exists():
        await message.answer_photo(FSInputFile(step.photo_path), caption=cap, reply_markup=markup)
    elif step.photo_file_id:
        await message.answer_photo(step.photo_file_id, caption=cap, reply_markup=markup)
    elif step.photo_url:
        await message.answer_photo(step.photo_url, caption=cap, reply_markup=markup)
    else:
        await message.answer(cap, reply_markup=markup)

# ---------- Хендлеры ----------
@router.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await ensure_db()
    await state.clear()

    if await has_participated(message.from_user.id):
        email = await get_user_email(message.from_user.id)
        await message.answer(
            "Вы уже проходили этот квиз. Спасибо за участие! ✅\n"
            f"Ваш e-mail: <b>{email or '—'}</b>"
        )
        return

    # deep-link payload: "/start <payload>"
    payload = ""
    parts = (message.text or "").strip().split(maxsplit=1)
    if len(parts) == 2:
        payload = parts[1].strip()

    # Проверка приглашения: разовый пропуск
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT 1 FROM access_users WHERE user_id=? LIMIT 1", (message.from_user.id,))
        allowed = (await cur.fetchone()) is not None

    if not allowed:
        if not INVITE_SECRET or payload != INVITE_SECRET:
            await message.answer("Этот бот доступен только по приглашению. 🔒\nПерейдите по специальной ссылке.")
            return
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute(
                "INSERT OR IGNORE INTO access_users(user_id, granted_at) VALUES(?, ?)",
                (message.from_user.id, int(time.time()))
            )
            await db.commit()

    await state.set_state(CollectEmail.wait_email)
    if INTRO_PHOTO.exists():
        await message.answer_photo(FSInputFile(str(INTRO_PHOTO)), caption=INTRO_TEXT)
    else:
        await message.answer(INTRO_TEXT)

@router.message(CollectEmail.wait_email)
async def handle_email(message: Message, state: FSMContext):
    # Защита от второго прохождения
    if await has_participated(message.from_user.id):
        email = await get_user_email(message.from_user.id)
        await message.answer("Этот квиз можно пройти только один раз. ✅\n"
                             f"Ранее вы указали e-mail: <b>{email or '—'}</b>")
        return

    raw = (message.text or "").strip()
    if not is_valid_email(raw):
        await message.answer("❌ Адрес выглядит некорректно. Введите e-mail ещё раз (пример: name@example.com).")
        return

    local, domain = raw.split("@", 1)
    try:
        domain_ascii = domain_to_ascii(domain)
    except Exception:
        await message.answer("❌ Домен e-mail некорректен (IDN). Проверьте написание и попробуйте снова.")
        return

    mx_ok = check_mx(domain_ascii)
    if not mx_ok:
        await message.answer("❌ Похоже, у домена нет MX-записей. Проверьте адрес и отправьте снова.")
        return

    email_norm = f"{local}@{domain_ascii}".lower()

    # Дубликаты (БД + Excel)
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT 1 FROM emails WHERE LOWER(email)=LOWER(?) LIMIT 1", (email_norm,))
        exists_db = (await cur.fetchone()) is not None
    exists_xlsx = False
    if EXCEL_PATH.exists():
        try:
            df = pd.read_excel(str(EXCEL_PATH))
            if 'email' in df.columns:
                exists_xlsx = email_norm in df['email'].astype(str).str.lower().values
        except Exception:
            pass
    if exists_db or exists_xlsx:
        await message.answer(
            f"❌ Email {email_norm} уже зарегистрирован.\n"
            "Если вы считаете это ошибкой — отправьте другой адрес или введите /restart."
        )
        return

    ok, msg = await save_email_row(message.from_user, email_norm, domain_ascii, mx_ok)
    if not ok:
        await message.answer(f"❌ {msg}")
        return

    try:
        append_excel_row(email_norm, domain_ascii, mx_ok, message.from_user)
    except Exception as e:
        print(f"Excel append error: {e}")

    await message.answer("✅ Спасибо! E-mail принят и успешно зарегистрирован!", reply_markup=ReplyKeyboardRemove())

    await state.set_state(Quiz.q)
    await state.update_data(q=0, score=0)
    await send_step(message, STEPS[0])

@router.message(Quiz.q)
async def handle_quiz_answer(message: Message, state: FSMContext):
    data = await state.get_data()
    idx = int(data.get("q", 0))
    score = int(data.get("score", 0))
    step = STEPS[idx]

    user_text = (message.text or "").strip()
    if user_text not in step.options:
        await message.answer("Пожалуйста, выберите один из вариантов на кнопках 🙂", reply_markup=kb(step.options))
        return

    if user_text == step.correct:
        score += 1
        await message.answer(step.feedback_ok, reply_markup=ReplyKeyboardRemove())
    else:
        await message.answer(step.feedback_bad, reply_markup=ReplyKeyboardRemove())

    idx += 1
    if idx < len(STEPS):
        await state.update_data(q=idx, score=score)
        await send_step(message, STEPS[idx])
    else:
        await state.clear()
        if FINAL_PHOTO.exists():
            await message.answer_photo(FSInputFile(str(FINAL_PHOTO)), caption=FINAL_TEXT)
        else:
            await message.answer(FINAL_TEXT)

# --- Служебные команды (для админов) ---
@router.message(Command("export"))
async def cmd_export(message: Message):
    if message.from_user.id not in settings.admin_ids:
        await message.answer("Команда доступна только администраторам.")
        return
    await ensure_db()
    total = await export_excel_from_db()
    await message.answer(f"✅ Excel перегенерирован. Строк: {total}\nФайл: {EXCEL_PATH}")

@router.message(Command("stats"))
async def cmd_stats(message: Message):
    if message.from_user.id not in settings.admin_ids:
        await message.answer("Команда доступна только администраторам.")
        return
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            cur = await db.execute("SELECT COUNT(*) FROM emails")
            db_count = (await cur.fetchone())[0]
        excel_count = 0
        if EXCEL_PATH.exists():
            df = pd.read_excel(str(EXCEL_PATH))
            excel_count = len(df)
        await message.answer(f"📊 Статистика email адресов:\n"
                             f"• В базе данных: {db_count}\n"
                             f"• В Excel файле: {excel_count}\n"
                             f"• Уникальных: {max(db_count, excel_count)}")
    except Exception as e:
        await message.answer(f"❌ Ошибка при получении статистики: {e}")

@router.message(Command("debug_images"))
async def cmd_debug_images(message: Message):
    if message.from_user.id not in settings.admin_ids:
        await message.answer("Команда доступна только администраторам.")
        return
    pairs = [
        ("INTRO_PHOTO", INTRO_PHOTO),
        ("FINAL_PHOTO", FINAL_PHOTO),
        ("q1", ASSETS_DIR / "q1.png"),
        ("q2", ASSETS_DIR / "q2.png"),
        ("q3", ASSETS_DIR / "q3.png"),
        ("q4", ASSETS_DIR / "q4.png"),
    ]
    lines = [f"{name}: {p} — {'OK' if Path(p).exists() else 'NOT FOUND'}" for name, p in pairs]
    await message.answer("\n".join(lines))

# --- Шлагбаум: пропускаем состояния и системные команды, остальным требуем инвайт ---
@router.message()
async def gate_all(message: Message, state: FSMContext):
    st = await state.get_state()
    if st in {CollectEmail.wait_email.state, Quiz.q.state}:
        return
    txt = (message.text or "").strip()
    allowed_cmds = ("/start", "/status", "/help", "/export", "/stats", "/debug_images")
    if any(txt.startswith(c) for c in allowed_cmds):
        return
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT 1 FROM access_users WHERE user_id=? LIMIT 1", (message.from_user.id,))
        allowed = (await cur.fetchone()) is not None
    if not allowed:
        await message.answer("Доступ по приглашению. Перейдите по специальной ссылке. 🔒")
        return

# ---------- Запуск ----------
async def main():
    if not settings.bot_token:
        raise SystemExit("Укажите TELEGRAM_TOKEN в .env")
    await ensure_db()
    bot = Bot(
        token=settings.bot_token,
        default=DefaultBotProperties(parse_mode=ParseMode.HTML),
    )
    dp = Dispatcher()
    dp.include_router(router)
    await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())

if __name__ == "__main__":
    asyncio.run(main())
